"""Excel 리포터.

openpyxl 기반 .xlsx 리포트 생성기.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from anshim.core.analyzers.hybrid import HybridScanResult
from anshim.core.reporters.base import BaseReporter, ReportData

logger = logging.getLogger(__name__)

# 심각도별 셀 배경색 (ARGB)
SEVERITY_FILL: dict[str, str] = {
    "critical": "FFfee2e2",
    "high": "FFffedd5",
    "medium": "FFfef9c3",
    "low": "FFdbeafe",
}

# 헤더 색상
HEADER_FILL = "FF1e3a5f"
HEADER_FONT_COLOR = "FFFFFFFF"


def _make_fill(argb: str):
    from openpyxl.styles import PatternFill
    return PatternFill(fill_type="solid", fgColor=argb)


def _make_header_style():
    from openpyxl.styles import Alignment, Font, PatternFill
    return (
        Font(bold=True, color=HEADER_FONT_COLOR),
        PatternFill(fill_type="solid", fgColor=HEADER_FILL),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
    )


def _auto_width(ws, min_width: int = 10, max_width: int = 60) -> None:
    """열 너비 자동 조정 (MergedCell 무시)."""
    from openpyxl.cell.cell import MergedCell

    for col in ws.columns:
        # MergedCell은 column_letter 속성이 없으므로 건너뜀
        first_real = next((c for c in col if not isinstance(c, MergedCell)), None)
        if first_real is None:
            continue
        col_letter = first_real.column_letter
        max_len = min_width
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


class ExcelReporter(BaseReporter):
    """openpyxl 기반 Excel 리포트 생성기."""

    def generate(self, scan_result: HybridScanResult, output_path: Path) -> Path:
        """Excel 리포트 파일 생성.

        Args:
            scan_result: 하이브리드 스캔 결과.
            output_path: 출력 파일 경로 (또는 디렉토리).

        Returns:
            생성된 .xlsx 파일 경로.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError as e:
            raise RuntimeError("openpyxl이 설치되지 않았습니다: pip install openpyxl") from e

        data = ReportData.from_hybrid_result(scan_result)

        # 출력 경로 결정
        if output_path.is_dir() or (not output_path.suffix):
            output_path.mkdir(parents=True, exist_ok=True)
            output_file = output_path / f"anshim_report_{data.scan_id}.xlsx"
        else:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_file = output_path

        wb = Workbook()

        self._write_summary_sheet(wb, data)
        self._write_vulns_sheet(wb, data)
        self._write_compliance_sheet(wb, data)

        # 기본 시트(Sheet) 제거
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(output_file))
        logger.info("Excel 리포트 생성: %s", output_file)
        return output_file

    def _write_summary_sheet(self, wb, data: ReportData) -> None:
        from openpyxl.styles import Alignment, Font
        ws = wb.create_sheet("요약")
        hfont, hfill, halign = _make_header_style()

        # 제목
        ws["A1"] = "AnShim 보안 감사 리포트"
        ws["A1"].font = Font(bold=True, size=16)
        ws.merge_cells("A1:C1")

        ws["A2"] = f"생성 일시: {data.generated_at}"
        ws["A2"].font = Font(italic=True, color="FF6b7280")
        ws.merge_cells("A2:C2")

        # 스캔 개요
        row = 4
        ws.cell(row=row, column=1, value="항목")
        ws.cell(row=row, column=2, value="값")
        for col in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign

        overview = [
            ("스캔 ID", data.scan_id),
            ("대상 경로", data.target_path),
            ("분석 파일", f"{data.scanned_files} / {data.total_files}"),
            ("컴플라이언스", ", ".join(data.compliance_types).upper()),
            ("분석 방식", "하이브리드 (규칙 + LLM)" if data.llm_enabled else "규칙 기반"),
            ("소요 시간", f"{data.duration_seconds:.1f}초"),
        ]
        for i, (label, value) in enumerate(overview, start=row + 1):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)

        # 취약점 통계
        row = row + len(overview) + 2
        ws.cell(row=row, column=1, value="심각도")
        ws.cell(row=row, column=2, value="건수")
        for col in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign

        stats = [
            ("Critical", data.critical_count, "FFdc2626"),
            ("High", data.high_count, "FFea580c"),
            ("Medium", data.medium_count, "FFca8a04"),
            ("Low", data.low_count, "FF2563eb"),
            ("전체", data.total_issues, "FF1f2937"),
        ]
        for i, (label, count, color) in enumerate(stats, start=row + 1):
            ws.cell(row=i, column=1, value=label)
            cell = ws.cell(row=i, column=2, value=count)
            cell.font = Font(bold=True, color=color)

        _auto_width(ws)

    def _write_vulns_sheet(self, wb, data: ReportData) -> None:
        from openpyxl.styles import Alignment, Font
        ws = wb.create_sheet("취약점 목록")
        hfont, hfill, halign = _make_header_style()

        headers = ["번호", "심각도", "파일 경로", "라인", "규칙 ID", "제목", "설명", "출처", "수정 제안"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign

        severity_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
        sorted_results = sorted(
            data.results,
            key=lambda r: severity_order.get(r.severity.lower(), 4),
        )

        for row_idx, result in enumerate(sorted_results, start=2):
            sev = result.severity.lower()
            fill = _make_fill(SEVERITY_FILL.get(sev, "FFFFFFFF"))

            # 수정 제안 텍스트 추출
            fix_text = ""
            if isinstance(result.remediation, dict):
                fix_text = result.remediation.get("fix_summary", "")
            elif result.remediation:
                fix_text = str(result.remediation)

            row_data = [
                row_idx - 1,
                result.severity.upper(),
                result.file_path,
                result.line_start,
                result.rule_id,
                result.title,
                result.description or "",
                result.source.upper(),
                fix_text,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 행 높이 설정
        for row in ws.iter_rows(min_row=2):
            ws.row_dimensions[row[0].row].height = 30

        _auto_width(ws)

    def _write_compliance_sheet(self, wb, data: ReportData) -> None:
        from openpyxl.styles import Alignment, Font
        ws = wb.create_sheet("ISMS-P 매핑")
        hfont, hfill, halign = _make_header_style()

        headers = ["컴플라이언스", "항목 ID", "항목 제목", "카테고리", "심각도", "파일 경로", "라인", "취약점 제목"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign

        row_idx = 2
        for result in data.results:
            for mapping in result.compliance_mappings:
                sev = result.severity.lower()
                fill = _make_fill(SEVERITY_FILL.get(sev, "FFFFFFFF"))

                row_data = [
                    mapping.compliance_type.upper(),
                    mapping.compliance_id,
                    mapping.compliance_title,
                    mapping.compliance_category,
                    result.severity.upper(),
                    result.file_path,
                    result.line_start,
                    result.title,
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.fill = fill
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

                row_idx += 1

        if row_idx == 2:
            ws.cell(row=2, column=1, value="컴플라이언스 매핑 데이터 없음")

        _auto_width(ws)
