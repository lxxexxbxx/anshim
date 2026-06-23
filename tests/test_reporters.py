"""리포터 테스트."""

import json
import tempfile
from pathlib import Path

import pytest

from anshim.core.analyzers.hybrid import HybridScanResult
from anshim.core.compliance.mapper import ComplianceMappingInfo, MappedResult
from anshim.core.reporters import ExcelReporter, HTMLReporter, JSONReporter, get_reporter


def _make_result(severity: str = "high", with_mapping: bool = True) -> MappedResult:
    mappings = []
    if with_mapping:
        mappings = [
            ComplianceMappingInfo(
                compliance_type="isms-p",
                compliance_id="2.7.1",
                compliance_title="암호화 적용",
                compliance_category="2.7 암호화 적용",
                rule_id="test-rule",
            ),
            ComplianceMappingInfo(
                compliance_type="owasp",
                compliance_id="A02:2021",
                compliance_title="Cryptographic Failures",
                compliance_category="OWASP Top 10",
                rule_id="test-rule",
            ),
        ]
    return MappedResult(
        rule_id="bandit.B303",
        title="취약한 암호화 알고리즘 사용",
        description="MD5는 취약한 해시 알고리즘입니다.",
        severity=severity,
        file_path="app/auth.py",
        line_start=42,
        line_end=42,
        code_snippet="hashlib.md5(password.encode()).hexdigest()",
        source="bandit",
        confidence="high",
        compliance_mappings=mappings,
    )


def _make_scan_result(num_results: int = 3) -> HybridScanResult:
    severities = ["critical", "high", "medium", "low"]
    results = [
        _make_result(severity=severities[i % len(severities)], with_mapping=(i % 2 == 0))
        for i in range(num_results)
    ]
    return HybridScanResult(
        scan_id="test1234",
        target_path="/tmp/test-project",
        total_files=10,
        scanned_files=8,
        duration_seconds=2.5,
        model_used=None,
        compliance_types=["isms-p", "owasp"],
        llm_enabled=False,
        results=results,
        total_issues=len(results),
        critical_count=sum(1 for r in results if r.severity == "critical"),
        high_count=sum(1 for r in results if r.severity == "high"),
        medium_count=sum(1 for r in results if r.severity == "medium"),
        low_count=sum(1 for r in results if r.severity == "low"),
        false_positives_removed=0,
        compliance_summary={
            "isms-p": {"total": 2, "by_severity": {"critical": 1, "medium": 1}},
            "owasp": {"total": 2, "by_severity": {"critical": 1, "medium": 1}},
        },
    )


class TestHTMLReporter:
    def test_generate_creates_file(self, tmp_path: Path) -> None:
        reporter = HTMLReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        assert output.exists()
        assert output.suffix == ".html"

    def test_html_contains_basic_structure(self, tmp_path: Path) -> None:
        reporter = HTMLReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        html = output.read_text(encoding="utf-8")
        assert "<!DOCTYPE html>" in html
        assert "AnShim" in html
        assert "취약점" in html
        assert "test1234" in html

    def test_html_contains_severity_counts(self, tmp_path: Path) -> None:
        reporter = HTMLReporter()
        scan = _make_scan_result(num_results=4)
        output = reporter.generate(scan, tmp_path)

        html = output.read_text(encoding="utf-8")
        assert "Critical" in html or "critical" in html.lower()
        assert "High" in html or "high" in html.lower()

    def test_html_contains_compliance_section(self, tmp_path: Path) -> None:
        reporter = HTMLReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        html = output.read_text(encoding="utf-8")
        assert "ISMS-P" in html or "isms-p" in html.lower()

    def test_empty_results(self, tmp_path: Path) -> None:
        reporter = HTMLReporter()
        scan = _make_scan_result(num_results=0)
        output = reporter.generate(scan, tmp_path)

        html = output.read_text(encoding="utf-8")
        assert "취약점이 발견되지 않았습니다" in html

    def test_output_to_specific_file(self, tmp_path: Path) -> None:
        reporter = HTMLReporter()
        scan = _make_scan_result()
        out_file = tmp_path / "my_report.html"
        output = reporter.generate(scan, out_file)

        assert output == out_file
        assert output.exists()


class TestExcelReporter:
    def test_generate_creates_file(self, tmp_path: Path) -> None:
        reporter = ExcelReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        assert output.exists()
        assert output.suffix == ".xlsx"

    def test_workbook_has_required_sheets(self, tmp_path: Path) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        reporter = ExcelReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        wb = load_workbook(str(output))
        assert "요약" in wb.sheetnames
        assert "취약점 목록" in wb.sheetnames
        assert "ISMS-P 매핑" in wb.sheetnames

    def test_vuln_sheet_has_data(self, tmp_path: Path) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        reporter = ExcelReporter()
        scan = _make_scan_result(num_results=2)
        output = reporter.generate(scan, tmp_path)

        wb = load_workbook(str(output))
        ws = wb["취약점 목록"]
        # 헤더 행 + 데이터 행
        assert ws.max_row >= 2

    def test_summary_sheet_has_headers(self, tmp_path: Path) -> None:
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        reporter = ExcelReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        wb = load_workbook(str(output))
        ws = wb["요약"]
        # A1에 AnShim 제목 포함
        assert ws["A1"].value and "AnShim" in str(ws["A1"].value)


class TestJSONReporter:
    def test_generate_creates_file(self, tmp_path: Path) -> None:
        reporter = JSONReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        assert output.exists()
        assert output.suffix == ".json"

    def test_json_is_valid(self, tmp_path: Path) -> None:
        reporter = JSONReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        data = json.loads(output.read_text(encoding="utf-8"))
        assert isinstance(data, dict)

    def test_json_has_required_fields(self, tmp_path: Path) -> None:
        reporter = JSONReporter()
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        data = json.loads(output.read_text(encoding="utf-8"))
        assert "scan_id" in data
        assert "summary" in data
        assert "statistics" in data
        assert "results" in data
        assert data["scan_id"] == "test1234"
        assert data["statistics"]["total_issues"] == 3

    def test_json_results_structure(self, tmp_path: Path) -> None:
        reporter = JSONReporter()
        scan = _make_scan_result(num_results=1)
        output = reporter.generate(scan, tmp_path)

        data = json.loads(output.read_text(encoding="utf-8"))
        assert len(data["results"]) == 1
        r = data["results"][0]
        assert "rule_id" in r
        assert "severity" in r
        assert "file_path" in r
        assert "compliance_mappings" in r

    def test_sarif_format(self, tmp_path: Path) -> None:
        reporter = JSONReporter(sarif=True)
        scan = _make_scan_result()
        output = reporter.generate(scan, tmp_path)

        assert output.suffix == ".sarif"
        data = json.loads(output.read_text(encoding="utf-8"))
        assert data["version"] == "2.1.0"
        assert "runs" in data
        assert len(data["runs"]) == 1
        assert "tool" in data["runs"][0]
        assert "results" in data["runs"][0]


class TestReporterFactory:
    def test_get_html_reporter(self) -> None:
        reporter = get_reporter("html")
        assert isinstance(reporter, HTMLReporter)

    def test_get_excel_reporter(self) -> None:
        reporter = get_reporter("excel")
        assert isinstance(reporter, ExcelReporter)

    def test_get_json_reporter(self) -> None:
        reporter = get_reporter("json")
        assert isinstance(reporter, JSONReporter)
        assert reporter.sarif is False

    def test_get_sarif_reporter(self) -> None:
        reporter = get_reporter("sarif")
        assert isinstance(reporter, JSONReporter)
        assert reporter.sarif is True

    def test_xlsx_alias(self) -> None:
        reporter = get_reporter("xlsx")
        assert isinstance(reporter, ExcelReporter)

    def test_case_insensitive(self) -> None:
        reporter = get_reporter("HTML")
        assert isinstance(reporter, HTMLReporter)

    def test_invalid_format_raises(self) -> None:
        with pytest.raises(ValueError, match="지원하지 않는 리포트 형식"):
            get_reporter("pdf")
